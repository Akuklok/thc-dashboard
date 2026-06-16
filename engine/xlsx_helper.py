"""
Small Excel writer that works on pandas 3.0 + openpyxl, where pandas'
own ExcelWriter raises "At least one sheet must be visible".

write_sheets(path, {"Sheet name": dataframe, ...})  -> writes a multi-sheet .xlsx
Sheet order is preserved. NaN becomes a blank cell.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def write_sheets(path, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        ws = wb.create_sheet(title=str(name)[:31])
        d = df.astype(object).where(pd.notna(df), None)
        for row in dataframe_to_rows(d, index=False, header=True):
            ws.append(row)
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    wb.active = 0
    wb.save(path)
