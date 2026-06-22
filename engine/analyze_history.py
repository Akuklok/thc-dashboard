"""
ANALYZE THC HISTORY  --  turns the daily history cache (thc_history.parquet)
into buyer-ready seasonality, trend, event-lift and velocity intelligence.

Outputs "THC History Insights.xlsx" with:
  Category Seasonality   - monthly demand index per category (1.00 = average month)
  Item Seasonality       - items with strong seasonal peaks + the month to BUY AHEAD
  Event Lifts            - 4/20, Green Wednesday, NYE, July 4 demand multipliers
  YoY Growth             - this year vs last year, by category and top items
  True Velocity          - real trailing 30/90-day units/day per item (from actual days)

Methodology notes:
  - Only FULL calendar months (>=25 days of data) are used for seasonality/YoY,
    so partial months never distort the pattern.
  - Seasonality index = (avg units in that calendar month) / (avg monthly units),
    averaged across all years available. >1.15 = seasonal peak.
"""
import os, datetime
import numpy as np
import pandas as pd
from xlsx_helper import write_sheets

CACHE = r"C:\Users\Anna K\OneDrive - Top Ten Liquors\Documents\Claude\thc_history.parquet"
OUT_FOLDERS = [r"C:\Users\Anna K\Downloads",
               r"C:\Users\Anna K\OneDrive - Top Ten Liquors\THC Reports"]
MIN_DAYS_IN_MONTH = 25     # a calendar month must have at least this many days to count
PEAK_THRESHOLD    = 1.15   # index above this = a real seasonal peak
MIN_ITEM_UNITS    = 200    # ignore tiny items for item-level seasonality

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def load():
    h = pd.read_parquet(CACHE)
    h["Date"] = pd.to_datetime(h["Date"])
    h["Y"] = h["Date"].dt.year
    h["M"] = h["Date"].dt.month
    h["YM"] = h["Date"].dt.to_period("M")
    h["Units"] = pd.to_numeric(h["Units"], errors="coerce").fillna(0)
    h["Revenue"] = pd.to_numeric(h["Revenue"], errors="coerce").fillna(0)
    return h

def full_months(h):
    """Set of YM periods that have >= MIN_DAYS_IN_MONTH days of data."""
    days = h.groupby("YM")["Date"].nunique()
    return set(days[days >= MIN_DAYS_IN_MONTH].index)

def category_seasonality(h, fm):
    g = h[h["YM"].isin(fm)]
    # monthly total units per category per (year,month)
    mt = g.groupby(["Category", "YM"])["Units"].sum().reset_index()
    mt["M"] = mt["YM"].dt.month
    rows = []
    for cat, d in mt.groupby("Category"):
        base = d["Units"].mean()
        if base <= 0:
            continue
        idx = d.groupby("M")["Units"].mean() / base
        row = {"Category": cat, "Avg Units/Mo": round(base)}
        for m in range(1, 13):
            row[MONTHS[m-1]] = round(idx.get(m, np.nan), 2) if m in idx.index else np.nan
        peaks = [MONTHS[m-1] for m in idx.index if idx[m] >= PEAK_THRESHOLD]
        row["Peak Months"] = ", ".join(peaks)
        rows.append(row)
    cols = ["Category", "Avg Units/Mo"] + MONTHS + ["Peak Months"]
    return pd.DataFrame(rows)[cols].sort_values("Avg Units/Mo", ascending=False)

def item_seasonality(h, fm):
    g = h[h["YM"].isin(fm)]
    tot = g.groupby("Product Code")["Units"].sum()
    keep = tot[tot >= MIN_ITEM_UNITS].index
    g = g[g["Product Code"].isin(keep)]
    mt = g.groupby(["Product Code", "YM"])["Units"].sum().reset_index()
    mt["M"] = mt["YM"].dt.month
    names = g.groupby("Product Code")["Product Description"].first()
    cats = g.groupby("Product Code")["Category"].first()
    rows = []
    for pc, d in mt.groupby("Product Code"):
        base = d["Units"].mean()
        if base <= 0 or d["M"].nunique() < 6:   # need at least half the year covered
            continue
        idx = d.groupby("M")["Units"].mean() / base
        peak_m = int(idx.idxmax()); peak_v = idx.max()
        if peak_v < PEAK_THRESHOLD:
            continue
        buy_ahead = MONTHS[(peak_m - 2) % 12]   # the month before the peak
        rows.append({"Item": names.get(pc), "Category": cats.get(pc),
                     "Annual Units": int(tot.get(pc, 0)),
                     "Peak Month": MONTHS[peak_m-1], "Peak Index": round(peak_v, 2),
                     "Buy-Ahead Month": buy_ahead})
    if not rows:
        return pd.DataFrame(columns=["Item","Category","Annual Units","Peak Month","Peak Index","Buy-Ahead Month"])
    return pd.DataFrame(rows).sort_values(["Peak Index","Annual Units"], ascending=False)

def item_monthly_index(h, fm):
    """Per-item monthly demand index (1.00 = that item's average month). Used by the
    brief for item-level season-aware buying. Only items with enough history."""
    g = h[h["YM"].isin(fm)]
    tot = g.groupby("Product Code")["Units"].sum()
    keep = tot[tot >= MIN_ITEM_UNITS].index
    g = g[g["Product Code"].isin(keep)]
    mt = g.groupby(["Product Code", "YM"])["Units"].sum().reset_index()
    mt["M"] = mt["YM"].dt.month
    names = g.groupby("Product Code")["Product Description"].first()
    cats = g.groupby("Product Code")["Category"].first()
    rows = []
    for pc, d in mt.groupby("Product Code"):
        base = d["Units"].mean()
        if base <= 0 or d["M"].nunique() < 6:   # need at least half the year covered
            continue
        idx = d.groupby("M")["Units"].mean() / base
        row = {"Product Code": pc, "Item": names.get(pc), "Category": cats.get(pc),
               "Avg Units/Mo": round(base)}
        for m in range(1, 13):
            row[MONTHS[m-1]] = round(idx.get(m), 2) if m in idx.index else np.nan
        rows.append(row)
    cols = ["Product Code", "Item", "Category", "Avg Units/Mo"] + MONTHS
    return pd.DataFrame(rows, columns=cols).sort_values("Avg Units/Mo", ascending=False)

def event_lifts(h):
    daily = h.groupby("Date")["Units"].sum().sort_index()
    def lift(dt):
        if dt not in daily.index:
            return None
        win = daily[(daily.index >= dt - pd.Timedelta(days=21)) &
                    (daily.index <= dt + pd.Timedelta(days=21)) &
                    (daily.index != dt)]
        base = win.median()
        return None if not base else (daily[dt], base, daily[dt] / base)
    def thanksgiving(y):
        d = datetime.date(y, 11, 1)
        thurs = [d + datetime.timedelta(days=i) for i in range(31)
                 if (d + datetime.timedelta(days=i)).weekday() == 3
                 and (d + datetime.timedelta(days=i)).month == 11]
        return thurs[3]  # 4th Thursday
    events = []
    yrs = sorted(h["Y"].unique())
    for y in yrs:
        events += [(f"4/20 {y}", pd.Timestamp(y,4,20)),
                   (f"July 4 {y}", pd.Timestamp(y,7,4)),
                   (f"NYE {y}", pd.Timestamp(y,12,31))]
        tg = thanksgiving(y)
        events.append((f"Green Wed {y}", pd.Timestamp(tg - datetime.timedelta(days=1))))
    rows = []
    for name, dt in events:
        r = lift(dt)
        if r:
            rows.append({"Event": name, "Date": dt.date(),
                         "Units That Day": int(r[0]), "Normal Day": int(r[1]),
                         "Lift x": round(r[2], 1)})
    return pd.DataFrame(rows).sort_values("Lift x", ascending=False) if rows else pd.DataFrame()

def yoy_growth(h, fm):
    g = h[h["YM"].isin(fm)]
    mt = g.groupby(["Category", "YM"])["Units"].sum().reset_index()
    mt["M"] = mt["YM"].dt.month; mt["Y"] = mt["YM"].dt.year
    rows = []
    for cat, d in mt.groupby("Category"):
        piv = d.pivot_table(index="M", columns="Y", values="Units")
        yrs = sorted(piv.columns)
        if len(yrs) < 2:
            continue
        prev, cur = yrs[-2], yrs[-1]
        common = piv[[prev, cur]].dropna()
        if common.empty:
            continue
        p, c = common[prev].sum(), common[cur].sum()
        if p <= 0:
            continue
        rows.append({"Category": cat, f"{prev} (shared mos)": int(p),
                     f"{cur} (shared mos)": int(c), "YoY %": round((c-p)/p*100, 1)})
    return pd.DataFrame(rows).sort_values("YoY %", ascending=False) if rows else pd.DataFrame()

def true_velocity(h):
    last = h["Date"].max()
    out = []
    for pc, d in h.groupby("Product Code"):
        d30 = d[d["Date"] > last - pd.Timedelta(days=30)]["Units"].sum()
        d90 = d[d["Date"] > last - pd.Timedelta(days=90)]["Units"].sum()
        out.append({"Item": d["Product Description"].iloc[-1],
                    "Category": d["Category"].iloc[-1],
                    "Units/Day 30D": round(d30/30, 2), "Units/Day 90D": round(d90/90, 2),
                    "Momentum (30D vs 90D)": round((d30/30)/(d90/90), 2) if d90 > 0 else np.nan})
    v = pd.DataFrame(out)
    return v.sort_values("Units/Day 90D", ascending=False)

def style(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
        for col in ws.columns:
            wd = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(wd+2, 10), 48)
    wb.save(path)

def main():
    if not os.path.exists(CACHE):
        raise SystemExit("No history cache yet - run build_history.py first.")
    h = load()
    fm = full_months(h)
    print(f"history: {h['Date'].nunique()} days, {h['Date'].min():%Y-%m-%d}->{h['Date'].max():%Y-%m-%d}, "
          f"{len(fm)} full months")
    cat = category_seasonality(h, fm)
    itm = item_seasonality(h, fm)
    imi = item_monthly_index(h, fm)
    evt = event_lifts(h)
    yoy = yoy_growth(h, fm)
    vel = true_velocity(h)
    for folder in OUT_FOLDERS:
        os.makedirs(folder, exist_ok=True)
        path = os.path.join(folder, "THC History Insights.xlsx")
        try:
            write_sheets(path, {"Category Seasonality": cat, "Item Seasonality": itm,
                                "Item Monthly Index": imi, "Event Lifts": evt,
                                "YoY Growth": yoy, "True Velocity": vel})
            style(path)
            print(f"wrote -> {path}")
        except PermissionError:
            print(f"  locked, skipped: {path}")
    print(f"\nSeasonal items found: {len(itm)} | Events measured: {len(evt)}")

if __name__ == "__main__":
    main()
