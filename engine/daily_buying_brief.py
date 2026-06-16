"""
THC DAILY BUYING BRIEF  --  reusable generator
================================================================================
Turns your latest exports into ONE ranked "do this today" brief — the content an
automatic buying notification would send. Channel-agnostic (email/Teams/SMS later).

INPUTS (newest matching file is auto-picked from the folders below):
  THC Transfer Sales Report*.xlsx   -> per-store on-hand + sales  (REQUIRED: restock/transfer/buy)
  Buncha_File_Export*.xlsx          -> competitor prices          (optional: pricing flags)
  Full Inventory Sales Report*.xlsx -> our prices                 (optional: pricing flags)
  THC <date>.xlsx (buyer sheet)     -> discount/margin/deal terms (optional: enriches buys)

OUTPUTS (written to each OUT_FOLDER; locked files are skipped):
  THC Daily Buying Brief.txt   -> the readable brief (notification body) + a dated copy
  THC Daily Buying Brief.xlsx  -> Brief / Restock & Transfer / Pricing Flags tabs + dated copy

RUN:  python daily_buying_brief.py   (or run_daily_brief.bat)
"""
import glob, os, re, datetime
import pandas as pd
import numpy as np
from xlsx_helper import write_sheets
from config_locations import drop_warehouses
from filepick import newest_readable, read_newest

# ----------------------- SETTINGS -----------------------
INPUT_FOLDERS = [r"C:\Users\Anna K\Downloads",
                 r"C:\Users\Anna K\OneDrive - Top Ten Liquors\Top Ten OneDrive - Reports - New Version",
                 r"C:\Users\Anna K\OneDrive - Top Ten Liquors\Documents",
                 r"C:\Users\Anna K\OneDrive - Top Ten Liquors\Documents\Claude"]
OUT_FOLDERS   = [r"C:\Users\Anna K\Downloads",
                 r"C:\Users\Anna K\OneDrive - Top Ten Liquors\THC Reports"]
TARGET_WEEKS  = 3      # weeks-of-supply target per store
ABOVE_FLAG    = 10     # flag if our price >= this % above the cheapest competitor
BELOW_FLAG    = 15     # flag if our price >= this % below the cheapest competitor
TOP_N         = 15     # actions shown in the text brief
COMPETITORS   = ["Total Wine", "Cub"]   # named competitors tracked in THC Competitor Prices.xlsx
MONTHS        = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
BUY_AHEAD_TOP = 8        # seasonal buy-ahead items shown in the text brief
# --------------------------------------------------------

def norm(u):
    s = re.sub(r"\D", "", str(u)); return s.lstrip("0") or s

def newest(pattern):
    return newest_readable(pattern, INPUT_FOLDERS)

def find_buyer_sheet():
    """Locate the buyer master file by structure (sheet 'THC' w/ 'Discount To Market')."""
    cands = []
    for f in INPUT_FOLDERS: cands += glob.glob(os.path.join(f, "THC*.xlsx"))
    for f in sorted(cands, key=os.path.getmtime, reverse=True):
        try:
            head = pd.read_excel(f, sheet_name="THC", header=None, nrows=3)
            if head.iloc[2].astype(str).str.contains("Discount To Market").any():
                return f
        except Exception:
            continue
    return None

def buyer_lookup():
    """upc -> (discount%, gm%, deal terms) from the buyer master sheet."""
    f = find_buyer_sheet()
    if not f: return {}, None
    d = pd.read_excel(f, sheet_name="THC", header=None).iloc[3:]
    out = {}
    for _, r in d.iterrows():
        upc = norm(r[8])
        if upc == "": continue
        disc = pd.to_numeric(r[26], errors="coerce")
        gm   = pd.to_numeric(r[29], errors="coerce")
        deal = str(r[16]).strip()
        out[upc] = (None if pd.isna(disc) else round(disc*100, 1),
                    None if pd.isna(gm) else round(gm*100, 1),
                    "" if deal.lower() == "nan" else deal)
    return out, os.path.basename(f)

def seasonal_factors(today):
    """Per-category demand multiplier heading INTO the coverage window:
    next-month seasonality index / this-month index (from history insights).
    >1 = demand rising into a peak (buy more); <1 = cooling (buy less)."""
    f = newest("THC History Insights.xlsx")
    out = {}
    if not f: return out
    try:
        c = pd.read_excel(f, sheet_name="Category Seasonality")
    except Exception:
        return out
    cur, nxt = MONTHS[today.month - 1], MONTHS[today.month % 12]
    for _, r in c.iterrows():
        try:
            a, b = float(r[cur]), float(r[nxt])
            if a > 0:
                out[str(r["Category"]).strip()] = min(max(b / a, 0.8), 1.5)
        except Exception:
            pass
    return out

def item_seasonal_factors(today):
    """ITEM-level demand multiplier into the coverage window (next-month index /
    this-month index, per product from history). Sharper than category; used first,
    with the category factor as fallback for items lacking enough history."""
    f = newest("THC History Insights.xlsx")
    out = {}
    if not f: return out
    try:
        d = pd.read_excel(f, sheet_name="Item Monthly Index")
    except Exception:
        return out
    cur, nxt = MONTHS[today.month - 1], MONTHS[today.month % 12]
    for _, r in d.iterrows():
        try:
            a, b = float(r[cur]), float(r[nxt])
            if a > 0 and pd.notna(b):
                out[norm(r["Product Code"])] = min(max(b / a, 0.7), 1.8)
        except Exception:
            pass
    return out

def restock(buyers, today):
    # Prefer the daily Full Inventory Sales Report (on-hand + velocity per store);
    # fall back to the transfer report if that's what's present.
    rd = lambda x: pd.read_excel(x, sheet_name=0)
    df, f = read_newest("Full Inventory Sales Report*.xlsx", INPUT_FOLDERS, rd)
    if df is None:
        df, f = read_newest("THC Transfer Sales Report*.xlsx", INPUT_FOLDERS, rd)
    if df is None: return None, None
    if "Department" in df.columns:   # Full Inventory has all depts -> keep THC only
        df = df[df["Department"].astype(str).str.contains("THC", case=False, na=False)].copy()
    df = drop_warehouses(df)   # stores only -- a warehouse is not a transfer point or a stockout
    for c in ("OH", "30D", "90D"):
        df[c] = pd.to_numeric(df.get(c), errors="coerce").fillna(0)
    df["price"] = pd.to_numeric(df["Price"], errors="coerce")
    # TREND-AWARE weekly velocity: weight the recent 30 days over the 90-day baseline,
    # so a store stocks up as demand accelerates and eases off as it cools.
    vel30, vel90 = df["30D"]*7/30, df["90D"]*7/90
    df["vel"]   = 0.6*vel30 + 0.4*vel90
    df["trend"] = np.where(vel90 > 0, vel30/vel90, 1.0)
    # SEASON-AWARE target: item-level seasonal factor first (sharper), category fallback.
    ISF = item_seasonal_factors(today)          # upc -> item factor
    SF  = seasonal_factors(today)               # category -> factor (fallback)
    cat_f = (df["Category"].astype(str).str.strip().map(SF)
             if "Category" in df.columns else pd.Series(np.nan, index=df.index))
    df["seasonF"] = df["Product Code"].map(norm).map(ISF).fillna(cat_f).fillna(1.0)
    df["seasonSrc"] = np.where(df["Product Code"].map(norm).map(ISF).notna(), "item", "category")
    df["target"]= np.ceil(df["vel"]*TARGET_WEEKS*df["seasonF"])
    df["OHp"]   = df["OH"].clip(lower=0)
    df["need"]  = (df["target"]-df["OHp"]).clip(lower=0)
    df["surp"]  = (df["OHp"]-df["target"]).clip(lower=0)
    code = df.groupby("Product Description")["Product Code"].first().to_dict()
    rows = []
    for item, g in df.groupby("Product Description"):
        price = g["price"].median()
        pool = dict(zip(g.loc[g.surp>0, "Location Name"], g.loc[g.surp>0, "surp"]))
        disc, gm, deal = buyers.get(norm(code.get(item, "")), (None, None, ""))
        for _, r in g[g.need>0].sort_values("vel", ascending=False).iterrows():
            need, got, src = r["need"], 0, []
            for loc in sorted(pool, key=pool.get, reverse=True):
                if need <= 0: break
                take = min(need, pool[loc])
                if take > 0: src.append((loc, int(take))); pool[loc]-=take; need-=take; got+=take
            buy = int(max(0, need))
            act = ("Transfer " + ", ".join(f"{t} from {l}" for l,t in src)) if src else ""
            if buy > 0: act = (act+"; " if act else "") + f"BUY {buy}"
            tr = r["trend"]
            trend = "rising" if tr >= 1.2 else "falling" if tr <= 0.8 else "steady"
            rows.append({"Item":item, "Store":r["Location Name"],
                         "Stockout?":"STOCKOUT" if (r["OHp"]==0 and r["vel"]>0) else "",
                         "On Hand":int(r["OH"]), "Wk Velocity":round(r["vel"],1),
                         "Trend":trend, "Season x":round(float(r["seasonF"]),2),
                         "Season Basis":r.get("seasonSrc",""),
                         "Need":int(r["need"]), "Transfer Qty":int(got), "Buy Qty":buy,
                         "Wk $ at Risk":round(r["vel"]*price,2) if pd.notna(price) else 0,
                         "Discount %":disc, "GM %":gm, "Deal Terms":deal, "Action":act})
    R = pd.DataFrame(rows).sort_values("Wk $ at Risk", ascending=False)
    return R, os.path.basename(f)

def pricing():
    """One pricing table combining NAMED competitors (Total Wine / Cub, from the
    competitor sheet) with the BLENDED Buncha market. For each item the benchmark is
    the cheapest named competitor if we have one, otherwise the Buncha market price.
    Keeps broad market coverage AND shows named-competitor detail where available."""
    I, inv = read_newest("Full Inventory Sales Report*.xlsx", INPUT_FOLDERS,
                         lambda x: pd.read_excel(x, sheet_name=0))
    if I is None: return None
    I = I[I["Department"].astype(str).str.contains("THC", case=False, na=False)]
    I["upc"] = I["Product Code"].map(norm); I["P"] = pd.to_numeric(I["Price"], errors="coerce")
    ours = (I.groupby(["upc", "Product Description"])["P"].median().reset_index()
              .rename(columns={"P": "Our Price", "Product Description": "Item"}))
    ours = ours[ours["Our Price"] > 0]

    # blended Buncha market (optional)
    bf = newest("Buncha_File_Export*.xlsx")
    if bf:
        B = pd.read_excel(bf, usecols=["item_upc", "original_price", "category"])
        B = B[B["category"].astype(str).str.upper() == "THC"]; B["upc"] = B["item_upc"].map(norm)
        B = (B[B["upc"] != ""].groupby("upc")["original_price"].mean()
               .rename("Buncha Mkt").reset_index())
        ours = ours.merge(B, on="upc", how="left")
    else:
        ours["Buncha Mkt"] = np.nan

    # named competitor prices (optional)
    present = []
    cf = newest("THC Competitor Prices.xlsx")
    if cf:
        try:
            C = pd.read_excel(cf, sheet_name="Competitor Prices")
            C["upc"] = C["Product Code"].map(norm)
            for c in COMPETITORS:
                col = f"{c} $"
                if col in C.columns:
                    C[col] = pd.to_numeric(C[col], errors="coerce"); present.append(c)
            ours = ours.merge(C[["upc"] + [f"{c} $" for c in present]], on="upc", how="left")
        except Exception:
            pass

    rows = []
    for _, r in ours.iterrows():
        named = {c: r[f"{c} $"] for c in present if pd.notna(r.get(f"{c} $"))}
        buncha = r.get("Buncha Mkt")
        if named:
            bench = min(named.values()); vs = "/".join(k for k, v in named.items() if v == bench)
        elif pd.notna(buncha):
            bench = buncha; vs = "Market"
        else:
            continue
        if not bench or bench <= 0:   # skip bad/zero benchmark prices (avoids divide-by-zero)
            continue
        gap = (r["Our Price"] - bench) / bench * 100
        flag = ("ABOVE market (review)" if gap >= ABOVE_FLAG
                else "BELOW market (raise)" if gap <= -BELOW_FLAG else "")
        if not flag: continue
        row = {"Item": r["Item"], "Our Price": round(float(r["Our Price"]), 2)}
        for c in COMPETITORS:
            v = r.get(f"{c} $")
            row[c] = round(float(v), 2) if pd.notna(v) else None
        row["Buncha Mkt"] = round(float(buncha), 2) if pd.notna(buncha) else None
        row["Market Price"] = round(float(bench), 2)
        row["Vs"] = vs; row["Gap %"] = round(gap, 1); row["Flag"] = flag
        rows.append(row)
    if not rows: return None
    cols = ["Item", "Our Price"] + COMPETITORS + ["Buncha Mkt", "Market Price", "Vs", "Gap %", "Flag"]
    return pd.DataFrame(rows)[cols].sort_values("Gap %", ascending=False)

def buy_ahead(today):
    """Items whose seasonal peak is NEXT month (i.e. buy-ahead month == this month),
    read from THC History Insights.xlsx. Returns None if insights aren't available."""
    f = newest("THC History Insights.xlsx")
    if not f:
        return None
    try:
        d = pd.read_excel(f, sheet_name="Item Seasonality")
    except Exception:
        return None
    cur = MONTHS[today.month - 1]
    if "Buy-Ahead Month" not in d.columns:
        return None
    d = d[d["Buy-Ahead Month"] == cur].copy()
    if d.empty:
        return None
    return d.sort_values("Annual Units", ascending=False)

def text_brief(R, P, BA, today):
    L = ["="*70, f"  THC DAILY BUYING BRIEF  -  {today:%A %b %d, %Y}", "="*70, ""]
    if R is not None:
        so = (R["Stockout?"]=="STOCKOUT").sum()
        tr = R["Action"].str.startswith("Transfer").sum()
        by = R["Action"].str.contains("BUY").sum()
        L += [f"* {so} stockouts on sellers | {tr} fixable by transfer | {by} need a buy", ""]
    if P is not None:
        L += [f"* {(P['Flag'].str.startswith('ABOVE')).sum()} priced ABOVE market (review) | "
              f"{(P['Flag'].str.startswith('BELOW')).sum()} BELOW market (raise)", ""]
    if BA is not None and len(BA):
        L += [f"* {len(BA)} seasonal items to BUY AHEAD now (they peak next month)", ""]
    if R is not None:
        L.append(f"TOP {TOP_N} ACTIONS (by weekly $ at risk):")
        for _, r in R.head(TOP_N).iterrows():
            tag = "STOCKOUT" if r["Stockout?"] else r.get("Trend", "low")
            ctx = f"  [disc {r['Discount %']}% gm {r['GM %']}%]" if (r["Buy Qty"]>0 and pd.notna(r["Discount %"])) else ""
            L.append(f"  ${r['Wk $ at Risk']:>6.0f}/wk [{tag:8}] {str(r['Item'])[:40]:40} @ {str(r['Store'])[:14]:14} -> {str(r['Action'])[:55]}{ctx}")
        L.append("")
    if BA is not None and len(BA):
        L.append("BUY AHEAD (seasonal - order extra now, these peak next month):")
        for _, r in BA.head(BUY_AHEAD_TOP).iterrows():
            L.append(f"  {str(r['Item'])[:44]:44} peaks {r['Peak Month']} x{r['Peak Index']}  ({int(r['Annual Units'])}/yr)")
        L.append("")
    if P is not None and len(P):
        L.append("PRICING - top above market (review):")
        for _, r in P[P.Flag.str.startswith("ABOVE")].head(5).iterrows():
            L.append(f"  +{r['Gap %']:>4.0f}%  {str(r['Item'])[:46]:46} ours ${r['Our Price']:.2f} vs mkt ${r['Market Price']:.2f}")
        L.append("PRICING - room to raise (below market):")
        for _, r in P[P.Flag.str.startswith("BELOW")].head(5).iterrows():
            L.append(f"  {r['Gap %']:>5.0f}%  {str(r['Item'])[:46]:46} ours ${r['Our Price']:.2f} vs mkt ${r['Market Price']:.2f}")
    return "\n".join(L)

def write_outputs(brief, R, P, BA, today):
    stamp = f"{today:%Y-%m-%d}"
    for folder in OUT_FOLDERS:
        os.makedirs(folder, exist_ok=True)
        for name, content in ((f"THC Daily Buying Brief.txt", brief),
                              (f"THC Daily Buying Brief {stamp}.txt", brief)):
            try:
                with open(os.path.join(folder, name), "w", encoding="utf-8") as fh: fh.write(content)
            except PermissionError:
                print(f"  txt locked, skipped: {os.path.join(folder,name)}")
        for name in (f"THC Daily Buying Brief.xlsx", f"THC Daily Buying Brief {stamp}.xlsx"):
            path = os.path.join(folder, name)
            try:
                sheets = {"Brief": pd.DataFrame({"Daily Buying Brief": [l for l in brief.split("\n")]})}
                if R is not None: sheets["Restock & Transfer"] = R
                if P is not None: sheets["Pricing Flags"] = P
                if BA is not None: sheets["Buy Ahead"] = BA
                write_sheets(path, sheets)
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill
                wb = load_workbook(path)
                for ws in wb.worksheets:
                    ws.freeze_panes = "A2"
                    if ws.title != "Brief":
                        ws.auto_filter.ref = ws.dimensions
                        for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
                    for col in ws.columns:
                        wd = max((len(str(c.value)) for c in col if c.value), default=10)
                        ws.column_dimensions[col[0].column_letter].width = min(max(wd+2,12), 70 if ws.title=="Brief" else 46)
                wb.save(path)
            except PermissionError:
                print(f"  xlsx locked, skipped: {path}")
        print(f"  wrote brief -> {folder}")

if __name__ == "__main__":
    today = datetime.date.today()
    print("Loading buyer discount/margin lookup...")
    buyers, bname = buyer_lookup()
    print(f"  {'using '+bname if bname else 'no buyer sheet found (buys won''t show discount/margin)'}")
    print("Building restock/transfer/buy actions...")
    R, rname = restock(buyers, today)
    if R is None:
        raise SystemExit("No 'THC Transfer Sales Report*.xlsx' found - cannot build the brief.")
    print(f"  using {rname} ({len(R)} store-needs)")
    print("Building pricing flags...")
    try:
        P = pricing()
    except Exception as e:
        P = None
        print(f"  pricing flags FAILED (skipped, brief still builds): {type(e).__name__}: {e}")
    print(f"  {'pricing flags: '+str(len(P)) if P is not None else 'skipped'}")
    try:
        BA = buy_ahead(today)
    except Exception as e:
        BA = None
        print(f"  buy-ahead FAILED (skipped): {type(e).__name__}: {e}")
    print(f"  {'buy-ahead items: '+str(len(BA)) if BA is not None else 'no buy-ahead items this month'}")
    brief = text_brief(R, P, BA, today)
    write_outputs(brief, R, P, BA, today)
    print("\n" + brief)
