"""
compare.py  --  match a buyer-uploaded spreadsheet against our catalog/stock by
UPC and produce an efficient side-by-side comparison (no dumping raw rows at the AI).

The server passes in our data (already read from the repo):
  catalog_rows : "{dept} - Full List.csv"  -> code, description, OUR COST (Net/Unit),
                 retail, club, deal, ranking, units/case   (the buyer's price master)
  stock_rows   : "{dept} Inventory.csv"     -> description -> on-hand, weeks-of-supply

IMPORTANT (per buyer note): OUR COST = Net/Unit (supplier/buyer cost). Retail/Club are
the CUSTOMER prices and are reported separately, never used as "cost".
"""
import io, re, csv


def _norm_txt(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _digits(s):
    return re.sub(r"\D", "", str(s or ""))


def _num(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    t = re.sub(r"[^0-9.\-]", "", str(s))
    try:
        return float(t) if t not in ("", "-", ".") else None
    except ValueError:
        return None


# ---- read the uploaded file into (headers, list-of-row-lists) ----
def parse_upload(file_bytes, filename):
    name = (filename or "").lower()
    rows = []
    if name.endswith(".csv") or name.endswith(".txt"):
        text = file_bytes.decode("utf-8", "replace")
        rows = [r for r in csv.reader(io.StringIO(text))]
    else:  # xlsx/xlsm
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
    # find the header row: first row with >=2 non-empty cells that aren't all numbers
    hi = 0
    for i, r in enumerate(rows[:8]):
        cells = [c for c in r if c not in (None, "")]
        if len(cells) >= 2 and not all(_num(c) is not None for c in cells):
            hi = i
            break
    headers = [str(c).strip() if c is not None else "" for c in rows[hi]]
    body = [r for r in rows[hi + 1:] if any(c not in (None, "") for c in r)]
    return headers, body


# ---- detect which columns hold UPC / price / qty / description ----
def detect_columns(headers, body):
    H = [_norm_txt(h) for h in headers]
    idx = {"upc": None, "price": None, "qty": None, "desc": None}

    def find(patterns):
        for i, h in enumerate(H):
            if any(re.search(p, h) for p in patterns):
                return i
        return None

    idx["upc"] = find([r"\bupc\b", r"gtin", r"product\s*code", r"item\s*code", r"^code$", r"barcode", r"sku"])
    idx["desc"] = find([r"descri", r"product", r"item\s*name", r"^item$", r"^name$", r"title"])
    idx["price"] = find([r"net.?unit", r"unit\s*cost", r"\bcost\b", r"\bprice\b", r"each", r"\$"])
    idx["qty"] = find([r"\bqty\b", r"quantit", r"cases", r"on.?hand", r"\bavail", r"count", r"units"])

    # fallback: detect a UPC column by content (mostly 11-14 digit values)
    if idx["upc"] is None and body:
        best, bestscore = None, 0
        for c in range(len(headers)):
            vals = [row[c] for row in body[:40] if c < len(row)]
            score = sum(1 for v in vals if 11 <= len(_digits(v)) <= 14)
            if score > bestscore:
                best, bestscore = c, score
        if bestscore >= max(3, len(body[:40]) // 3):
            idx["upc"] = best
    return idx


def build_catalog(catalog_rows):
    """code -> our catalog facts. catalog_rows = list of dicts (Full List)."""
    cat = {}
    by_desc = {}
    for r in catalog_rows:
        code = _digits(r.get("Product UPC"))
        if not code:
            continue
        entry = {
            "item": r.get("Product Description"),
            "cost": _num(r.get("Net/Unit")),                 # OUR COST (buyer/supplier)
            "retail": _num(r.get("Retail")),                 # customer price
            "club": _num(r.get("Club")),                     # customer club price
            "deal": (r.get("Deal Description") or "").strip(),
            "rank": r.get("Sales Ranking"),
            "units_case": _num(r.get("Units/Case")),
            "supplier": (r.get("Supplier") or "").strip(),
        }
        cat[code] = entry
        d = _norm_txt(r.get("Product Description"))
        if d:
            by_desc[d] = code
    return cat, by_desc


def build_stock(stock_rows):
    """description -> live stock (Inventory.csv has no code, keyed by Item)."""
    st = {}
    for r in stock_rows:
        d = _norm_txt(r.get("Item"))
        if d:
            st[d] = {"oh": _num(r.get("Chain OH")), "wos": _num(r.get("WOS")),
                     "vel": _num(r.get("Wk Velocity"))}
    return st


def compare(file_bytes, filename, catalog_rows, stock_rows):
    headers, body = parse_upload(file_bytes, filename)
    if not headers:
        return {"error": "Could not read that file."}
    idx = detect_columns(headers, body)
    if idx["upc"] is None and idx["desc"] is None:
        return {"error": "Couldn't find a UPC/product-code or description column to match on.",
                "headers": headers}

    cat, by_desc = build_catalog(catalog_rows)
    stock = build_stock(stock_rows)

    def cell(row, i):
        return row[i] if (i is not None and i < len(row)) else None

    out = []
    matched = cheaper = low = over = notcarried = 0
    for row in body:
        their_code = _digits(cell(row, idx["upc"])) if idx["upc"] is not None else ""
        their_desc = cell(row, idx["desc"])
        their_price = _num(cell(row, idx["price"]))
        their_qty = _num(cell(row, idx["qty"]))

        c = cat.get(their_code)
        if not c and their_desc:                      # name fallback
            code2 = by_desc.get(_norm_txt(their_desc))
            c = cat.get(code2) if code2 else None
            if code2:
                their_code = code2
        carry = c is not None
        rec = {"their_item": (their_desc or (c["item"] if c else "") or ""),
               "upc": their_code, "carry": "Yes" if carry else "No",
               "their_price": their_price, "their_qty": their_qty}
        if carry:
            matched += 1
            st = stock.get(_norm_txt(c["item"]), {})
            rec.update({"our_cost": c["cost"], "our_retail": c["retail"], "our_club": c["club"],
                        "oh": st.get("oh"), "wos": st.get("wos"), "deal": c["deal"], "rank": c["rank"]})
            if their_price is not None and c["cost"]:
                rec["vs_cost_pct"] = round((c["cost"] - their_price) / c["cost"] * 100, 1)
                if rec["vs_cost_pct"] > 0:
                    cheaper += 1
            wos = st.get("wos")
            if wos is not None and wos < 2:
                rec["flag"] = "LOW — restock"; low += 1
            elif wos is not None and wos > 12:
                rec["flag"] = "OVERSTOCK — skip"; over += 1
        else:
            notcarried += 1
            rec["flag"] = "we don't carry / not found"
        out.append(rec)

    summary = {"total": len(out), "matched": matched, "not_carried": notcarried,
               "cheaper_than_our_cost": cheaper, "low_stock": low, "overstock": over,
               "detected": {k: (headers[v] if v is not None else None) for k, v in idx.items()}}
    cols = ["their_item", "upc", "carry", "their_price", "our_cost", "vs_cost_pct",
            "our_retail", "our_club", "oh", "wos", "deal", "flag"]
    return {"summary": summary, "cols": cols, "rows": out}
